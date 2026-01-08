"""CSV and Excel import service for bulk data imports."""

import csv
import io
import logging
import uuid
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.driver import Driver
from app.models.equipment import Equipment
from app.models.load import Load, LoadStop
from app.models.worker import Worker, WorkerRole, WorkerStatus, WorkerType
from app.schemas.imports import (
    DriverImportRow,
    EquipmentImportRow,
    ImportError,
    ImportResult,
    LoadImportRow,
    ValidationResult,
)

logger = logging.getLogger(__name__)


class ImportService:
    """Service for handling CSV imports."""

    # Required columns for each entity type
    REQUIRED_COLUMNS = {
        "drivers": ["first_name", "last_name", "email", "phone"],
        "equipment": ["unit_number", "equipment_type"],
        "loads": [
            "customer_name",
            "pickup_city",
            "pickup_state",
            "pickup_zip",
            "delivery_city",
            "delivery_state",
            "delivery_zip",
        ],
    }

    # Field name variations - maps alternative column names to standard names
    FIELD_MAPPINGS = {
        "drivers": {
            # Name variations
            "firstname": "first_name",
            "first": "first_name",
            "fname": "first_name",
            "driver_first_name": "first_name",
            "driver_firstname": "first_name",
            "lastname": "last_name",
            "last": "last_name",
            "lname": "last_name",
            "driver_last_name": "last_name",
            "driver_lastname": "last_name",
            "name": "first_name",  # If only "name" field, map to first_name
            "full_name": "first_name",  # If full_name, map to first_name (user may need to split)
            "driver_name": "first_name",
            # License/CDL number variations
            "cdl_number": "license_number",
            "cdl": "license_number",
            "cdl_no": "license_number",
            "license_no": "license_number",
            "driver_license": "license_number",
            "dl_number": "license_number",
            # License state variations
            "cdl_state": "license_state",
            "license_state_code": "license_state",
            "dl_state": "license_state",
            # License expiry variations
            "cdl_expiration": "license_expiry",
            "cdl_exp": "license_expiry",
            "license_expiration": "license_expiry",
            "license_exp": "license_expiry",
            "dl_expiry": "license_expiry",
            "dl_expiration": "license_expiry",
            # Employment variations
            "emp_type": "employment_type",
            "employee_type": "employment_type",
            "worker_type": "employment_type",
            # Pay variations
            "rate": "pay_rate",
            "compensation_rate": "pay_rate",
            "pay_per_mile": "pay_rate",
            "rate_type": "pay_type",
            "compensation_type": "pay_type",
            # Date variations
            "start_date": "hire_date",
            "employment_start": "hire_date",
            "hired_date": "hire_date",
            # Contact variations
            "phone_number": "phone",
            "mobile": "phone",
            "cell": "phone",
            "telephone": "phone",
            "contact_number": "phone",
            "driver_phone": "phone",
            "email_address": "email",
            "driver_email": "email",
            "contact_email": "email",
            # Address variations
            "street": "address",
            "address_line_1": "address",
            "street_address": "address",
            "zip_code": "zip",
            "postal_code": "zip",
            "zipcode": "zip",
        },
        "equipment": {
            # Equipment type variations
            "type": "equipment_type",
            "vehicle_type": "equipment_type",
            "asset_type": "equipment_type",
            # Unit number variations
            "unit": "unit_number",
            "unit_no": "unit_number",
            "truck_number": "unit_number",
            "trailer_number": "unit_number",
            "asset_number": "unit_number",
            # VIN variations
            "vin_number": "vin",
            "vehicle_identification_number": "vin",
            # Mileage variations
            "mileage": "current_mileage",
            "odometer": "current_mileage",
            "miles": "current_mileage",
            # GPS variations
            "gps_vendor": "gps_provider",
            "telematics_provider": "gps_provider",
            "gps_id": "gps_device_id",
            "device_id": "gps_device_id",
            "telematics_id": "gps_device_id",
        },
        "loads": {
            # Customer variations
            "customer": "customer_name",
            "client": "customer_name",
            "shipper": "customer_name",
            # Pickup variations
            "origin_city": "pickup_city",
            "pickup_location_city": "pickup_city",
            "origin_state": "pickup_state",
            "pickup_location_state": "pickup_state",
            "origin_zip": "pickup_zip",
            "pickup_postal_code": "pickup_zip",
            "pickup_date_time": "pickup_date",
            "scheduled_pickup": "pickup_date",
            "pickup_appointment": "pickup_date",
            # Delivery variations
            "destination_city": "delivery_city",
            "delivery_location_city": "delivery_city",
            "destination_state": "delivery_state",
            "delivery_location_state": "delivery_state",
            "destination_zip": "delivery_zip",
            "delivery_postal_code": "delivery_zip",
            "delivery_date_time": "delivery_date",
            "scheduled_delivery": "delivery_date",
            "delivery_appointment": "delivery_date",
            # Rate variations
            "rate": "base_rate",
            "freight_rate": "base_rate",
            "line_haul": "base_rate",
            "price": "base_rate",
            # Weight variations
            "total_weight": "weight",
            "gross_weight": "weight",
            # Reference variations
            "reference": "reference_number",
            "ref_number": "reference_number",
            "ref_no": "reference_number",
            "load_number": "reference_number",
            "pro_number": "reference_number",
            # Notes variations
            "notes": "special_instructions",
            "instructions": "special_instructions",
            "comments": "special_instructions",
            "remarks": "special_instructions",
        },
    }

    # Schema mapping
    SCHEMAS = {
        "drivers": DriverImportRow,
        "equipment": EquipmentImportRow,
        "loads": LoadImportRow,
    }

    def __init__(self, db: AsyncSession):
        self.db = db

    def validate_import(
        self, file_bytes: bytes, filename: str, entity_type: str
    ) -> tuple[int, int, List[ImportError], List[str], List[Dict[str, Any]]]:
        """
        Validate import file without actually importing.

        Returns: (total_rows, valid_rows, errors, warnings, sample_data)
        """
        try:
            # Parse file
            rows = self.parse_csv(file_bytes, filename, entity_type=entity_type)

            if not rows:
                return 0, 0, [], ["File is empty"], []

            # Validate headers
            headers = list(rows[0].keys())
            logger.info(f"DEBUG: Parsed headers from {filename}: {headers}")
            validation = self.validate_headers(headers, entity_type)

            if not validation.valid:
                errors = [ImportError(row=0, error=err) for err in validation.errors]
                return len(rows), 0, errors, validation.warnings, []

            # Validate each row
            schema = self.SCHEMAS[entity_type]
            errors = []
            warnings = []
            valid_count = 0

            for idx, row in enumerate(rows, start=1):
                is_valid, validation_errors, validated_data = self.validate_row(row, schema)

                if not is_valid:
                    errors.append(
                        ImportError(
                            row=idx,
                            error="; ".join(validation_errors),
                        )
                    )
                else:
                    valid_count += 1

            # Get sample data (first 3 valid rows)
            sample_data = []
            for row in rows[:3]:
                is_valid, _, validated_data = self.validate_row(row, schema)
                if is_valid:
                    sample_data.append(dict(validated_data))

            return len(rows), valid_count, errors, warnings, sample_data

        except Exception as e:
            return 0, 0, [ImportError(row=0, error=str(e))], [], []

    def _normalize_field_name(self, field_name: str, entity_type: str) -> str:
        """
        Normalize field name using field mappings.

        Example: 'cdl_number' -> 'license_number' for drivers
        """
        # Strip BOM character if present (common in Excel exports)
        field_name = field_name.lstrip('\ufeff')
        field_lower = field_name.lower().strip().replace(" ", "_")
        mappings = self.FIELD_MAPPINGS.get(entity_type, {})
        return mappings.get(field_lower, field_lower)

    def _parse_date(self, date_str: Optional[str]) -> Optional[date]:
        """Parse date string in various formats."""
        if not date_str or not date_str.strip():
            return None

        date_str = date_str.strip()
        formats = [
            "%Y-%m-%d",        # 2024-01-15
            "%m/%d/%Y",        # 01/15/2024
            "%d/%m/%Y",        # 15/01/2024
            "%Y/%m/%d",        # 2024/01/15
            "%m-%d-%Y",        # 01-15-2024
            "%d-%m-%Y",        # 15-01-2024
        ]

        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue

        return None

    def _parse_excel(self, file_bytes: bytes, filename: str, entity_type: str = None) -> List[Dict[str, Any]]:
        """Parse Excel file and return list of row dictionaries with normalized field names."""
        try:
            # Load workbook from bytes
            workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            sheet = workbook.active

            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())

            # Parse data rows
            rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                cleaned_row = {}
                for i, value in enumerate(row):
                    if i >= len(headers):
                        break
                    if not headers[i]:
                        continue

                    # Clean the field name
                    field_name = headers[i].lower().replace(" ", "_")
                    # Normalize using field mappings if entity_type provided
                    if entity_type:
                        field_name = self._normalize_field_name(field_name, entity_type)

                    # Clean the value - convert to string and strip
                    if value is not None:
                        # Handle date/datetime objects
                        if isinstance(value, (date, datetime)):
                            cleaned_row[field_name] = value.strftime("%Y-%m-%d")
                        else:
                            cleaned_row[field_name] = str(value).strip()
                    else:
                        cleaned_row[field_name] = None

                if cleaned_row:  # Only add non-empty rows
                    rows.append(cleaned_row)

            workbook.close()
            return rows

        except Exception as e:
            raise ValueError(f"Failed to parse Excel: {str(e)}")

    def parse_csv(self, file_bytes: bytes, filename: str, entity_type: str = None) -> List[Dict[str, Any]]:
        """Parse CSV or Excel file and return list of row dictionaries with normalized field names."""
        # Determine file type from filename
        filename_lower = filename.lower()
        if filename_lower.endswith((".xlsx", ".xls")):
            return self._parse_excel(file_bytes, filename, entity_type)

        # Otherwise parse as CSV
        try:
            # Detect encoding - try utf-8-sig first to handle BOM
            try:
                content = file_bytes.decode("utf-8-sig")  # Handles BOM automatically
            except UnicodeDecodeError:
                try:
                    content = file_bytes.decode("utf-8")
                except UnicodeDecodeError:
                    content = file_bytes.decode("latin-1")  # Fallback

            # Parse CSV
            csv_file = io.StringIO(content)
            reader = csv.DictReader(csv_file)

            # Convert to list and normalize field names
            rows = []
            for row in reader:
                cleaned_row = {}
                for k, v in row.items():
                    if not k:
                        continue
                    # Clean the field name - strip BOM if present
                    field_name = k.lstrip('\ufeff').strip().lower().replace(" ", "_")
                    # Normalize using field mappings if entity_type provided
                    if entity_type:
                        field_name = self._normalize_field_name(field_name, entity_type)
                    # Clean the value
                    cleaned_row[field_name] = v.strip() if v else None
                rows.append(cleaned_row)

            return rows

        except Exception as e:
            raise ValueError(f"Failed to parse file: {str(e)}")

    def validate_headers(
        self, headers: List[str], entity_type: str
    ) -> ValidationResult:
        """Validate that required headers are present (after normalization)."""
        required = self.REQUIRED_COLUMNS.get(entity_type, [])
        # Normalize headers using field mappings
        headers_normalized = [self._normalize_field_name(h, entity_type) for h in headers]

        errors = []
        warnings = []
        suggestions = []

        # Check for missing required columns
        missing_columns = []
        for req_col in required:
            if req_col not in headers_normalized:
                missing_columns.append(req_col)

                # Check for similar column names (typos)
                for header in headers_normalized:
                    if self._is_similar(req_col, header):
                        suggestions.append(
                            f"Did you mean '{req_col}' instead of '{header}'?"
                        )

        if missing_columns:
            errors.append(
                f"Missing required columns: {', '.join(missing_columns)}. "
                f"Found columns: {', '.join(headers_normalized[:10])}{'...' if len(headers_normalized) > 10 else ''}"
            )

        return ValidationResult(
            valid=len(errors) == 0,
            errors=errors,
            warnings=warnings,
            suggestions=suggestions,
        )

    def _is_similar(self, str1: str, str2: str, threshold: float = 0.7) -> bool:
        """Check if two strings are similar (simple similarity check)."""
        # Simple Levenshtein-like check
        if len(str1) < 3 or len(str2) < 3:
            return False

        # Check if one contains most characters of the other
        common_chars = sum(1 for c in str1 if c in str2)
        similarity = common_chars / max(len(str1), len(str2))
        return similarity >= threshold

    def validate_row(
        self, row: Dict[str, Any], schema: Any
    ) -> Tuple[bool, List[str], Optional[Any]]:
        """
        Validate a single row against a Pydantic schema.

        Returns:
            (is_valid, errors, validated_data)
        """
        try:
            # Filter out None values and empty strings
            cleaned_row = {k: v for k, v in row.items() if v not in [None, "", "null"]}
            validated = schema(**cleaned_row)
            return True, [], validated
        except ValidationError as e:
            errors = [f"{err['loc'][0]}: {err['msg']}" for err in e.errors()]
            return False, errors, None
        except Exception as e:
            return False, [str(e)], None

    def generate_template(self, entity_type: str) -> bytes:
        """Generate CSV template with headers and sample row."""
        if entity_type == "drivers":
            headers = [
                "first_name",
                "last_name",
                "email",
                "phone",
                "license_number",
                "license_state",
                "license_expiry",
                "hire_date",
                "employment_type",
                "pay_rate",
                "pay_type",
                "address",
                "city",
                "state",
                "zip",
            ]
            sample = [
                "John",
                "Doe",
                "john.doe@example.com",
                "555-123-4567",
                "D1234567",
                "CA",
                "2025-12-31",
                "2024-01-15",
                "FULL_TIME",
                "0.45",
                "per_mile",
                "123 Main St",
                "Los Angeles",
                "CA",
                "90001",
            ]
        elif entity_type == "equipment":
            headers = [
                "unit_number",
                "equipment_type",
                "status",
                "make",
                "model",
                "year",
                "vin",
                "current_mileage",
                "gps_provider",
                "gps_device_id",
            ]
            sample = [
                "T001",
                "TRUCK",
                "ACTIVE",
                "Freightliner",
                "Cascadia",
                "2022",
                "1FUJGHDV1JLNA1234",
                "50000",
                "SAMSARA",
                "DEV123",
            ]
        elif entity_type == "loads":
            headers = [
                "customer_name",
                "pickup_city",
                "pickup_state",
                "pickup_zip",
                "pickup_date",
                "pickup_time",
                "delivery_city",
                "delivery_state",
                "delivery_zip",
                "delivery_date",
                "delivery_time",
                "commodity",
                "weight",
                "base_rate",
                "reference_number",
                "special_instructions",
            ]
            sample = [
                "ABC Corp",
                "Chicago",
                "IL",
                "60601",
                "2024-02-01",
                "09:00",
                "Dallas",
                "TX",
                "75201",
                "2024-02-03",
                "17:00",
                "Electronics",
                "25000",
                "2500.00",
                "REF001",
                "Fragile - handle with care",
            ]
        else:
            raise ValueError(f"Unknown entity type: {entity_type}")

        # Generate CSV
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerow(sample)

        return output.getvalue().encode("utf-8")

    async def check_duplicate_driver(
        self, company_id: str, email: str, phone: str, license_number: Optional[str]
    ) -> Optional[str]:
        """Check if driver already exists. Returns driver ID if found."""
        result = await self.db.execute(
            select(Driver).where(
                Driver.company_id == company_id,
                (Driver.email == email)
                | (Driver.phone == phone)
                | (
                    (Driver.cdl_number == license_number)  # Use cdl_number field
                    if license_number
                    else False
                ),
            )
        )
        existing = result.scalar_one_or_none()
        return existing.id if existing else None

    async def check_duplicate_equipment(
        self, company_id: str, unit_number: str, vin: Optional[str]
    ) -> Optional[str]:
        """Check if equipment already exists. Returns equipment ID if found."""
        result = await self.db.execute(
            select(Equipment).where(
                Equipment.company_id == company_id,
                (Equipment.unit_number == unit_number)
                | ((Equipment.vin == vin) if vin else False),
            )
        )
        existing = result.scalar_one_or_none()
        return existing.id if existing else None

    async def import_drivers(
        self, company_id: str, file_bytes: bytes, filename: str
    ) -> ImportResult:
        """Import drivers from CSV file."""
        try:
            # Parse CSV with field name normalization
            rows = self.parse_csv(file_bytes, filename, entity_type="drivers")

            if not rows:
                return ImportResult(
                    total=0,
                    successful=0,
                    failed=0,
                    errors=[],
                    warnings=["CSV file is empty"],
                )

            # Validate headers
            headers = list(rows[0].keys())
            validation = self.validate_headers(headers, "drivers")

            if not validation.valid:
                return ImportResult(
                    total=0,
                    successful=0,
                    failed=0,
                    errors=[
                        ImportError(row=0, error=err) for err in validation.errors
                    ],
                    warnings=validation.warnings,
                )

            # Import rows
            successful_ids = []
            errors = []
            warnings = []
            schema = self.SCHEMAS["drivers"]

            for idx, row in enumerate(rows, start=1):
                try:
                    # Validate row
                    is_valid, validation_errors, validated_data = self.validate_row(
                        row, schema
                    )

                    if not is_valid:
                        errors.append(
                            ImportError(
                                row=idx,
                                error="; ".join(validation_errors),
                            )
                        )
                        continue

                    # Check for duplicates
                    duplicate_id = await self.check_duplicate_driver(
                        company_id,
                        validated_data.email,
                        validated_data.phone,
                        validated_data.license_number,
                    )

                    if duplicate_id:
                        warnings.append(
                            f"Row {idx}: Driver with email '{validated_data.email}' already exists (skipped)"
                        )
                        continue

                    # Parse dates
                    license_expiry = self._parse_date(validated_data.license_expiry)
                    hire_date = self._parse_date(validated_data.hire_date)

                    # Create Worker record first (drivers are workers in the payroll system)
                    worker_id = str(uuid.uuid4())
                    worker = Worker(
                        id=worker_id,
                        company_id=company_id,
                        type=WorkerType.EMPLOYEE,
                        role=WorkerRole.DRIVER,
                        first_name=validated_data.first_name,
                        last_name=validated_data.last_name,
                        email=validated_data.email.lower(),
                        phone=validated_data.phone,
                        status=WorkerStatus.ACTIVE,
                    )
                    self.db.add(worker)
                    await self.db.flush()

                    # Create driver record linked to worker
                    driver_id = str(uuid.uuid4())
                    driver_metadata = {}
                    if validated_data.pay_rate:
                        driver_metadata["pay_rate"] = validated_data.pay_rate
                    if validated_data.pay_type:
                        driver_metadata["pay_type"] = validated_data.pay_type
                    if hire_date:
                        driver_metadata["hire_date"] = hire_date.isoformat()
                    if validated_data.address:
                        driver_metadata["address"] = validated_data.address
                    if validated_data.city:
                        driver_metadata["city"] = validated_data.city
                    if validated_data.state:
                        driver_metadata["state"] = validated_data.state
                    if validated_data.zip:
                        driver_metadata["zip"] = validated_data.zip

                    # Store license_state and employment_type in metadata since they're not direct fields
                    if validated_data.license_state:
                        driver_metadata["license_state"] = validated_data.license_state
                    if validated_data.employment_type:
                        driver_metadata["employment_type"] = validated_data.employment_type

                    driver = Driver(
                        id=driver_id,
                        company_id=company_id,
                        worker_id=worker_id,  # Link to worker
                        first_name=validated_data.first_name,
                        last_name=validated_data.last_name,
                        email=validated_data.email,
                        phone=validated_data.phone,
                        cdl_number=validated_data.license_number,  # Map license_number to cdl_number
                        cdl_expiration=license_expiry,
                        profile_metadata=driver_metadata if driver_metadata else None,
                    )

                    self.db.add(driver)
                    await self.db.flush()
                    successful_ids.append(driver.id)

                except Exception as e:
                    errors.append(ImportError(row=idx, error=str(e)))

            # Commit all changes
            await self.db.commit()

            return ImportResult(
                total=len(rows),
                successful=len(successful_ids),
                failed=len(errors),
                errors=errors,
                created_ids=successful_ids,
                warnings=warnings,
            )

        except Exception as e:
            await self.db.rollback()
            logger.exception("Driver import failed")
            raise

    async def import_equipment(
        self, company_id: str, file_bytes: bytes, filename: str
    ) -> ImportResult:
        """Import equipment from CSV file."""
        try:
            # Parse CSV with field name normalization
            rows = self.parse_csv(file_bytes, filename, entity_type="equipment")

            if not rows:
                return ImportResult(
                    total=0,
                    successful=0,
                    failed=0,
                    errors=[],
                    warnings=["CSV file is empty"],
                )

            # Validate headers
            headers = list(rows[0].keys())
            validation = self.validate_headers(headers, "equipment")

            if not validation.valid:
                return ImportResult(
                    total=0,
                    successful=0,
                    failed=0,
                    errors=[
                        ImportError(row=0, error=err) for err in validation.errors
                    ],
                    warnings=validation.warnings,
                )

            # Import rows
            successful_ids = []
            errors = []
            warnings = []
            schema = self.SCHEMAS["equipment"]

            for idx, row in enumerate(rows, start=1):
                try:
                    # Validate row
                    is_valid, validation_errors, validated_data = self.validate_row(
                        row, schema
                    )

                    if not is_valid:
                        errors.append(
                            ImportError(
                                row=idx,
                                error="; ".join(validation_errors),
                            )
                        )
                        continue

                    # Check for duplicates
                    duplicate_id = await self.check_duplicate_equipment(
                        company_id,
                        validated_data.unit_number,
                        validated_data.vin,
                    )

                    if duplicate_id:
                        warnings.append(
                            f"Row {idx}: Equipment with unit number '{validated_data.unit_number}' already exists (skipped)"
                        )
                        continue

                    # Create equipment
                    equipment = Equipment(
                        id=str(uuid.uuid4()),
                        company_id=company_id,
                        unit_number=validated_data.unit_number,
                        equipment_type=validated_data.equipment_type.upper(),
                        status=validated_data.status or "ACTIVE",
                        make=validated_data.make,
                        model=validated_data.model,
                        year=validated_data.year,
                        vin=validated_data.vin,
                        current_mileage=validated_data.current_mileage,
                        gps_provider=validated_data.gps_provider,
                        gps_device_id=validated_data.gps_device_id,
                    )

                    self.db.add(equipment)
                    await self.db.flush()
                    successful_ids.append(equipment.id)

                except Exception as e:
                    errors.append(ImportError(row=idx, error=str(e)))

            # Commit all changes
            await self.db.commit()

            return ImportResult(
                total=len(rows),
                successful=len(successful_ids),
                failed=len(errors),
                errors=errors,
                created_ids=successful_ids,
                warnings=warnings,
            )

        except Exception as e:
            await self.db.rollback()
            logger.exception("Equipment import failed")
            raise

    async def import_loads(
        self, company_id: str, file_bytes: bytes, filename: str
    ) -> ImportResult:
        """
        Import loads from CSV file.

        Note: This is a basic implementation. Full load creation would require:
        - Customer lookup/creation
        - Location geocoding
        - Date/time parsing
        - More complex validation
        """
        try:
            # Parse CSV with field name normalization
            rows = self.parse_csv(file_bytes, filename, entity_type="loads")

            if not rows:
                return ImportResult(
                    total=0,
                    successful=0,
                    failed=0,
                    errors=[],
                    warnings=["CSV file is empty"],
                )

            # Validate headers
            headers = list(rows[0].keys())
            validation = self.validate_headers(headers, "loads")

            if not validation.valid:
                return ImportResult(
                    total=0,
                    successful=0,
                    failed=0,
                    errors=[
                        ImportError(row=0, error=err) for err in validation.errors
                    ],
                    warnings=validation.warnings,
                )

            # Import rows
            successful_ids = []
            errors = []
            warnings = []
            schema = self.SCHEMAS["loads"]

            for idx, row in enumerate(rows, start=1):
                try:
                    # Validate row
                    is_valid, validation_errors, validated_data = self.validate_row(
                        row, schema
                    )

                    if not is_valid:
                        errors.append(
                            ImportError(
                                row=idx,
                                error="; ".join(validation_errors),
                            )
                        )
                        continue

                    # Parse dates
                    pickup_date = self._parse_date(validated_data.pickup_date)
                    delivery_date = self._parse_date(validated_data.delivery_date)

                    # Create load record
                    load_id = str(uuid.uuid4())
                    load_metadata = {}
                    if validated_data.reference_number:
                        load_metadata["reference_number"] = validated_data.reference_number
                    if validated_data.weight:
                        load_metadata["weight"] = validated_data.weight

                    load = Load(
                        id=load_id,
                        company_id=company_id,
                        customer_name=validated_data.customer_name,
                        load_type="FTL",  # Default to Full Truckload
                        commodity=validated_data.commodity,
                        base_rate=validated_data.base_rate,
                        notes=validated_data.special_instructions,
                        status="PLANNED",
                        metadata_json=load_metadata if load_metadata else None,
                    )
                    self.db.add(load)
                    await self.db.flush()

                    # Create pickup stop
                    pickup_stop_id = str(uuid.uuid4())
                    pickup_scheduled_at = None
                    if pickup_date and validated_data.pickup_time:
                        try:
                            # Parse time (HH:MM format)
                            time_parts = validated_data.pickup_time.strip().split(":")
                            if len(time_parts) == 2:
                                hour = int(time_parts[0])
                                minute = int(time_parts[1])
                                pickup_scheduled_at = datetime.combine(
                                    pickup_date, datetime.min.time().replace(hour=hour, minute=minute)
                                )
                        except (ValueError, AttributeError):
                            pass  # Silently skip invalid time format
                    elif pickup_date:
                        pickup_scheduled_at = datetime.combine(pickup_date, datetime.min.time())

                    pickup_stop = LoadStop(
                        id=pickup_stop_id,
                        load_id=load.id,
                        sequence=1,
                        stop_type="PICKUP",
                        city=validated_data.pickup_city,
                        state=validated_data.pickup_state,
                        postal_code=validated_data.pickup_zip,
                        scheduled_at=pickup_scheduled_at,
                    )
                    self.db.add(pickup_stop)

                    # Create delivery stop
                    delivery_stop_id = str(uuid.uuid4())
                    delivery_scheduled_at = None
                    if delivery_date and validated_data.delivery_time:
                        try:
                            # Parse time (HH:MM format)
                            time_parts = validated_data.delivery_time.strip().split(":")
                            if len(time_parts) == 2:
                                hour = int(time_parts[0])
                                minute = int(time_parts[1])
                                delivery_scheduled_at = datetime.combine(
                                    delivery_date, datetime.min.time().replace(hour=hour, minute=minute)
                                )
                        except (ValueError, AttributeError):
                            pass  # Silently skip invalid time format
                    elif delivery_date:
                        delivery_scheduled_at = datetime.combine(delivery_date, datetime.min.time())

                    delivery_stop = LoadStop(
                        id=delivery_stop_id,
                        load_id=load.id,
                        sequence=2,
                        stop_type="DELIVERY",
                        city=validated_data.delivery_city,
                        state=validated_data.delivery_state,
                        postal_code=validated_data.delivery_zip,
                        scheduled_at=delivery_scheduled_at,
                    )
                    self.db.add(delivery_stop)

                    await self.db.flush()
                    successful_ids.append(load.id)

                except Exception as e:
                    errors.append(ImportError(row=idx, error=str(e)))

            # Commit all changes
            await self.db.commit()

            return ImportResult(
                total=len(rows),
                successful=len(successful_ids),
                failed=len(errors),
                errors=errors,
                created_ids=successful_ids,
                warnings=warnings,
            )

        except Exception as e:
            await self.db.rollback()
            logger.exception("Load import failed")
            raise
